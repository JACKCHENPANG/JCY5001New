import os
import io
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Tuple
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

class ExcelTemplateGenerator:
    """Excel模板生成器"""
    
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def create_battery_template(self) -> io.BytesIO:
        """创建电池出厂数据模板"""
        wb = Workbook()
        ws = wb.active
        ws.title = "电池出厂数据"
        
        # 设置列标题
        headers = [
            "电池ID*", "批次号", "电池类型", "标称容量(mAh)", 
            "标称电压(V)", "制造商", "生产日期(YYYY-MM-DD)", "备注"
        ]
        
        # 写入标题行
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # 设置列宽
        column_widths = [15, 15, 12, 15, 12, 15, 20, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # 添加数据验证
        # 电池类型验证
        cell_type_validation = DataValidation(
            type="list",
            formula1='"LFP,NMC,LCO,NCA,LTO"',
            showDropDown=True
        )
        cell_type_validation.error = "请选择有效的电池类型"
        cell_type_validation.errorTitle = "无效输入"
        ws.add_data_validation(cell_type_validation)
        cell_type_validation.add(f"C2:C1000")
        
        # 添加示例数据
        example_data = [
            ["BAT001", "BATCH001", "LFP", 3000, 3.2, "CATL", "2025-01-01", "示例数据"],
            ["BAT002", "BATCH001", "NMC", 2500, 3.7, "BYD", "2025-01-02", ""],
            ["BAT003", "BATCH002", "LCO", 2000, 3.6, "Panasonic", "2025-01-03", ""]
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col in [4, 5]:  # 数值列居中
                    cell.alignment = self.center_alignment
        
        # 添加说明工作表
        ws_info = wb.create_sheet("使用说明")
        instructions = [
            ["电池出厂数据导入模板使用说明", ""],
            ["", ""],
            ["必填字段：", ""],
            ["• 电池ID", "唯一标识符，不能重复"],
            ["", ""],
            ["可选字段：", ""],
            ["• 批次号", "电池生产批次"],
            ["• 电池类型", "LFP/NMC/LCO/NCA/LTO"],
            ["• 标称容量", "单位：mAh"],
            ["• 标称电压", "单位：V"],
            ["• 制造商", "电池制造商名称"],
            ["• 生产日期", "格式：YYYY-MM-DD"],
            ["• 备注", "其他说明信息"],
            ["", ""],
            ["注意事项：", ""],
            ["1. 请勿修改表头", ""],
            ["2. 电池ID必须唯一", ""],
            ["3. 日期格式必须正确", ""],
            ["4. 数值字段请输入有效数字", ""],
            ["5. 删除示例数据后再导入", ""]
        ]
        
        for row, (col1, col2) in enumerate(instructions, 1):
            ws_info.cell(row=row, column=1, value=col1)
            ws_info.cell(row=row, column=2, value=col2)
            if row == 1:
                ws_info.cell(row=row, column=1).font = Font(bold=True, size=14)
        
        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 30
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def create_test_results_template(self) -> io.BytesIO:
        """创建测试结果数据模板"""
        wb = Workbook()
        ws = wb.active
        ws.title = "测试结果数据"
        
        # 设置列标题
        headers = [
            "测试ID*", "批次ID*", "测试时间*", "电池ID", "通道号",
            "电压(V)", "Rs值(Ω)", "Rct值(Ω)", "容量(mAh)", 
            "厚度(mm)", "温度(°C)", "测试结果", "错误代码"
        ]
        
        # 写入标题行
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # 设置列宽
        column_widths = [15, 15, 20, 15, 10, 12, 12, 12, 12, 12, 12, 12, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # 添加数据验证
        # 测试结果验证
        result_validation = DataValidation(
            type="list",
            formula1='"pass,fail"',
            showDropDown=True
        )
        result_validation.error = "请选择 pass 或 fail"
        result_validation.errorTitle = "无效输入"
        ws.add_data_validation(result_validation)
        result_validation.add(f"L2:L1000")
        
        # 添加示例数据
        example_data = [
            ["TEST001", "BATCH001", "2025-06-08T10:00:00Z", "BAT001", 1, 
             3.7, 0.05, 0.02, 2980, 5.2, 25.0, "pass", ""],
            ["TEST002", "BATCH001", "2025-06-08T10:01:00Z", "BAT002", 2, 
             3.6, 0.06, 0.025, 2450, 5.1, 25.5, "pass", ""],
            ["TEST003", "BATCH001", "2025-06-08T10:02:00Z", "BAT003", 3, 
             3.5, 0.08, 0.03, 1950, 5.0, 26.0, "fail", "E001"]
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col in [5, 6, 7, 8, 9, 10, 11]:  # 数值列居中
                    cell.alignment = self.center_alignment
        
        # 添加说明工作表
        ws_info = wb.create_sheet("使用说明")
        instructions = [
            ["测试结果数据导入模板使用说明", ""],
            ["", ""],
            ["必填字段：", ""],
            ["• 测试ID", "唯一标识符"],
            ["• 批次ID", "对应的测试批次ID"],
            ["• 测试时间", "ISO格式：YYYY-MM-DDTHH:MM:SSZ"],
            ["", ""],
            ["可选字段：", ""],
            ["• 电池ID", "对应的电池ID"],
            ["• 通道号", "测试设备通道"],
            ["• 电压", "测试电压值(V)"],
            ["• Rs值", "串联电阻(Ω)"],
            ["• Rct值", "电荷转移电阻(Ω)"],
            ["• 容量", "测试容量(mAh)"],
            ["• 厚度", "电池厚度(mm)"],
            ["• 温度", "测试温度(°C)"],
            ["• 测试结果", "pass/fail"],
            ["• 错误代码", "失败时的错误代码"],
            ["", ""],
            ["注意事项：", ""],
            ["1. 测试ID必须唯一", ""],
            ["2. 批次ID必须存在", ""],
            ["3. 时间格式必须正确", ""],
            ["4. 数值字段请输入有效数字", ""],
            ["5. 删除示例数据后再导入", ""]
        ]
        
        for row, (col1, col2) in enumerate(instructions, 1):
            ws_info.cell(row=row, column=1, value=col1)
            ws_info.cell(row=row, column=2, value=col2)
            if row == 1:
                ws_info.cell(row=row, column=1).font = Font(bold=True, size=14)
        
        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 35
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def create_device_template(self) -> io.BytesIO:
        """创建设备信息模板"""
        wb = Workbook()
        ws = wb.active
        ws.title = "设备信息"
        
        # 设置列标题
        headers = [
            "设备ID*", "设备名称*", "设备型号", "固件版本", "状态", "备注"
        ]
        
        # 写入标题行
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # 设置列宽
        column_widths = [15, 20, 15, 15, 12, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # 添加数据验证
        # 状态验证
        status_validation = DataValidation(
            type="list",
            formula1='"active,inactive,maintenance"',
            showDropDown=True
        )
        status_validation.error = "请选择有效的设备状态"
        status_validation.errorTitle = "无效输入"
        ws.add_data_validation(status_validation)
        status_validation.add(f"E2:E1000")
        
        # 添加示例数据
        example_data = [
            ["JCY5001", "JCY5001测试设备", "JCY5001", "1.0.0", "active", "主要测试设备"],
            ["JCY5002", "JCY5001备用设备", "JCY5001", "1.0.1", "inactive", "备用设备"],
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col == 5:  # 状态列居中
                    cell.alignment = self.center_alignment
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

class ExcelDataParser:
    """Excel数据解析器"""
    
    def __init__(self):
        self.errors = []
        self.warnings = []
    
    def parse_battery_data(self, file_content: bytes) -> Tuple[List[Dict], List[str]]:
        """解析电池出厂数据"""
        self.errors = []
        self.warnings = []
        
        try:
            # 读取Excel文件
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=0)
            
            # 检查必需列
            required_columns = ["电池ID"]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                self.errors.append(f"缺少必需列: {', '.join(missing_columns)}")
                return [], self.errors
            
            # 重命名列以匹配数据库字段
            column_mapping = {
                "电池ID": "battery_id",
                "批次号": "batch_number",
                "电池类型": "cell_type",
                "标称容量(mAh)": "nominal_capacity",
                "标称电压(V)": "nominal_voltage",
                "制造商": "manufacturer",
                "生产日期(YYYY-MM-DD)": "production_date",
                "备注": "notes"
            }
            
            # 只保留存在的列
            existing_mapping = {k: v for k, v in column_mapping.items() if k in df.columns}
            df = df.rename(columns=existing_mapping)
            
            # 删除空行
            df = df.dropna(subset=['battery_id'])
            
            batteries = []
            for index, row in df.iterrows():
                battery_data = {}
                row_errors = []
                
                # 验证电池ID
                battery_id = str(row['battery_id']).strip()
                if not battery_id or battery_id.lower() == 'nan':
                    row_errors.append(f"第{index+2}行：电池ID不能为空")
                    continue
                
                if len(battery_id) < 3 or len(battery_id) > 50:
                    row_errors.append(f"第{index+2}行：电池ID长度必须在3-50个字符之间")
                
                battery_data['battery_id'] = battery_id
                
                # 处理可选字段
                if 'batch_number' in row and pd.notna(row['batch_number']):
                    battery_data['batch_number'] = str(row['batch_number']).strip()
                
                if 'cell_type' in row and pd.notna(row['cell_type']):
                    cell_type = str(row['cell_type']).strip().upper()
                    if cell_type in ['LFP', 'NMC', 'LCO', 'NCA', 'LTO']:
                        battery_data['cell_type'] = cell_type
                    else:
                        row_errors.append(f"第{index+2}行：无效的电池类型 {cell_type}")
                
                if 'nominal_capacity' in row and pd.notna(row['nominal_capacity']):
                    try:
                        capacity = float(row['nominal_capacity'])
                        if capacity > 0:
                            battery_data['nominal_capacity'] = capacity
                        else:
                            row_errors.append(f"第{index+2}行：标称容量必须大于0")
                    except (ValueError, TypeError):
                        row_errors.append(f"第{index+2}行：标称容量必须是有效数字")
                
                if 'nominal_voltage' in row and pd.notna(row['nominal_voltage']):
                    try:
                        voltage = float(row['nominal_voltage'])
                        if voltage > 0:
                            battery_data['nominal_voltage'] = voltage
                        else:
                            row_errors.append(f"第{index+2}行：标称电压必须大于0")
                    except (ValueError, TypeError):
                        row_errors.append(f"第{index+2}行：标称电压必须是有效数字")
                
                if 'manufacturer' in row and pd.notna(row['manufacturer']):
                    battery_data['manufacturer'] = str(row['manufacturer']).strip()
                
                if 'production_date' in row and pd.notna(row['production_date']):
                    try:
                        if isinstance(row['production_date'], str):
                            production_date = datetime.strptime(row['production_date'], '%Y-%m-%d').date()
                        else:
                            production_date = row['production_date'].date()
                        battery_data['production_date'] = production_date.isoformat()
                    except (ValueError, AttributeError):
                        row_errors.append(f"第{index+2}行：生产日期格式错误，请使用YYYY-MM-DD格式")
                
                if row_errors:
                    self.errors.extend(row_errors)
                else:
                    batteries.append(battery_data)
            
            return batteries, self.errors
            
        except Exception as e:
            self.errors.append(f"文件解析错误: {str(e)}")
            return [], self.errors
    
    def parse_test_results_data(self, file_content: bytes) -> Tuple[List[Dict], List[str]]:
        """解析测试结果数据"""
        self.errors = []
        self.warnings = []
        
        try:
            # 读取Excel文件
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=0)
            
            # 检查必需列
            required_columns = ["测试ID", "批次ID", "测试时间"]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                self.errors.append(f"缺少必需列: {', '.join(missing_columns)}")
                return [], self.errors
            
            # 重命名列以匹配数据库字段
            column_mapping = {
                "测试ID": "test_id",
                "批次ID": "batch_id",
                "测试时间": "test_time",
                "电池ID": "battery_id_str",
                "通道号": "channel_number",
                "电压(V)": "voltage",
                "Rs值(Ω)": "rs_value",
                "Rct值(Ω)": "rct_value",
                "容量(mAh)": "capacity",
                "厚度(mm)": "thickness",
                "温度(°C)": "temperature",
                "测试结果": "test_result",
                "错误代码": "error_code"
            }
            
            # 只保留存在的列
            existing_mapping = {k: v for k, v in column_mapping.items() if k in df.columns}
            df = df.rename(columns=existing_mapping)
            
            # 删除空行
            df = df.dropna(subset=['test_id'])
            
            test_results = []
            for index, row in df.iterrows():
                result_data = {}
                row_errors = []
                
                # 验证必需字段
                test_id = str(row['test_id']).strip()
                if not test_id or test_id.lower() == 'nan':
                    row_errors.append(f"第{index+2}行：测试ID不能为空")
                    continue
                result_data['test_id'] = test_id
                
                # 批次ID
                if pd.isna(row['batch_id']):
                    row_errors.append(f"第{index+2}行：批次ID不能为空")
                    continue
                try:
                    result_data['batch_id'] = int(row['batch_id'])
                except (ValueError, TypeError):
                    row_errors.append(f"第{index+2}行：批次ID必须是有效数字")
                    continue
                
                # 测试时间
                if pd.isna(row['test_time']):
                    row_errors.append(f"第{index+2}行：测试时间不能为空")
                    continue
                
                try:
                    if isinstance(row['test_time'], str):
                        # 尝试解析ISO格式时间
                        test_time = datetime.fromisoformat(row['test_time'].replace('Z', '+00:00'))
                    else:
                        test_time = row['test_time']
                    result_data['test_time'] = test_time.isoformat()
                except (ValueError, AttributeError):
                    row_errors.append(f"第{index+2}行：测试时间格式错误")
                    continue
                
                # 处理可选字段
                optional_fields = [
                    ('battery_id_str', str),
                    ('channel_number', int),
                    ('voltage', float),
                    ('rs_value', float),
                    ('rct_value', float),
                    ('capacity', float),
                    ('thickness', float),
                    ('temperature', float),
                    ('error_code', str)
                ]
                
                for field, field_type in optional_fields:
                    if field in row and pd.notna(row[field]):
                        try:
                            if field_type == str:
                                result_data[field] = str(row[field]).strip()
                            else:
                                value = field_type(row[field])
                                if field_type == float and value < 0:
                                    row_errors.append(f"第{index+2}行：{field} 不能为负数")
                                else:
                                    result_data[field] = value
                        except (ValueError, TypeError):
                            row_errors.append(f"第{index+2}行：{field} 格式错误")
                
                # 验证测试结果
                if 'test_result' in row and pd.notna(row['test_result']):
                    test_result = str(row['test_result']).strip().lower()
                    if test_result in ['pass', 'fail']:
                        result_data['test_result'] = test_result
                    else:
                        row_errors.append(f"第{index+2}行：测试结果必须是 pass 或 fail")
                
                if row_errors:
                    self.errors.extend(row_errors)
                else:
                    test_results.append(result_data)
            
            return test_results, self.errors
            
        except Exception as e:
            self.errors.append(f"文件解析错误: {str(e)}")
            return [], self.errors

class ExcelExporter:
    """Excel数据导出器"""
    
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_test_results(self, test_results: List[Dict], include_details: bool = False) -> io.BytesIO:
        """导出测试结果数据"""
        wb = Workbook()
        ws = wb.active
        ws.title = "测试结果"
        
        # 基础列
        headers = [
            "测试ID", "批次ID", "测试时间", "电池ID", "通道号",
            "电压(V)", "Rs值(Ω)", "Rct值(Ω)", "容量(mAh)", 
            "厚度(mm)", "温度(°C)", "测试结果", "错误代码"
        ]
        
        # 写入标题行
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # 写入数据
        for row, result in enumerate(test_results, 2):
            data = [
                result.get('test_id', ''),
                result.get('batch_id', ''),
                result.get('test_time', ''),
                result.get('battery_id', ''),
                result.get('channel_number', ''),
                result.get('voltage', ''),
                result.get('rs_value', ''),
                result.get('rct_value', ''),
                result.get('capacity', ''),
                result.get('thickness', ''),
                result.get('temperature', ''),
                result.get('test_result', ''),
                result.get('error_code', '')
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_statistics_report(self, statistics: Dict) -> io.BytesIO:
        """导出统计报告"""
        wb = Workbook()
        ws = wb.active
        ws.title = "统计报告"
        
        # 写入统计信息
        ws.cell(row=1, column=1, value="电池阻抗测试统计报告").font = Font(bold=True, size=16)
        ws.cell(row=2, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        row = 4
        ws.cell(row=row, column=1, value="总体统计").font = Font(bold=True, size=14)
        row += 1
        
        overall_stats = [
            ("总测试数", statistics.get('total_tests', 0)),
            ("通过测试数", statistics.get('pass_tests', 0)),
            ("失败测试数", statistics.get('fail_tests', 0)),
            ("通过率", f"{statistics.get('pass_rate', 0):.2f}%")
        ]
        
        for label, value in overall_stats:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # 每日统计
        if 'daily_stats' in statistics and statistics['daily_stats']:
            row += 2
            ws.cell(row=row, column=1, value="每日统计").font = Font(bold=True, size=14)
            row += 1
            
            # 表头
            daily_headers = ["日期", "总测试数", "通过数", "失败数", "通过率"]
            for col, header in enumerate(daily_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.border
            row += 1
            
            # 数据
            for daily_stat in statistics['daily_stats']:
                data = [
                    daily_stat.get('date', ''),
                    daily_stat.get('total', 0),
                    daily_stat.get('pass_count', 0),
                    daily_stat.get('fail_count', 0),
                    f"{daily_stat.get('pass_rate', 0):.2f}%"
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.border
                row += 1
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

