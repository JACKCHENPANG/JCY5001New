"""
Web前端专用API接口
为JCY5001AS Web平台提供增强的API服务
"""

from flask import Blueprint, jsonify, request, current_app, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from sqlalchemy import and_, or_, desc, func, text
from datetime import datetime, timedelta
import json
import io
import csv
import tempfile
import os

from models.user import User, Device, Battery, TestBatch, TestResult, ImpedanceDetail, db

# 尝试导入Excel支持
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    current_app.logger.warning("openpyxl未安装，Excel导出功能不可用")

web_api_bp = Blueprint('web_api', __name__)

@web_api_bp.route('/dashboard/stats', methods=['GET'])
@jwt_required()
def get_dashboard_stats():
    """获取仪表板统计数据"""
    try:
        current_user_id = get_jwt_identity()
        current_app.logger.info(f"获取仪表板数据，用户ID: {current_user_id}")

        user = User.query.get(current_user_id)
        if not user:
            current_app.logger.error(f"用户不存在: {current_user_id}")
            return jsonify({'error': 'User not found'}), 404

        current_app.logger.info(f"用户信息: {user.username}, 角色: {user.role}")

        # 根据用户角色过滤数据
        if user.role == 'admin':
            # 管理员可以看到所有数据
            device_query = Device.query
            result_query = TestResult.query
            current_app.logger.info("管理员用户，查询所有数据")
        else:
            # 普通用户只能看到自己的数据
            user_devices = Device.query.filter_by(user_id=current_user_id).all()
            device_ids = [d.id for d in user_devices]
            current_app.logger.info(f"普通用户，设备数量: {len(user_devices)}, 设备IDs: {device_ids}")

            device_query = Device.query.filter_by(user_id=current_user_id)
            # 修复查询：通过TestBatch正确关联到Device
            if device_ids:
                result_query = TestResult.query.join(TestBatch).filter(TestBatch.device_id.in_(device_ids))
            else:
                # 如果用户没有设备，返回空查询
                result_query = TestResult.query.filter(TestResult.id == -1)  # 永远不会匹配的条件

        # 设备统计
        total_devices = device_query.count()
        online_devices = device_query.filter_by(status='online').count()
        current_app.logger.info(f"设备统计: 总数={total_devices}, 在线={online_devices}")

        # 测试统计
        total_tests = result_query.count()
        today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_tests = result_query.filter(TestResult.test_start_time >= today_start).count()
        current_app.logger.info(f"测试统计: 总数={total_tests}, 今日={today_tests}")

        # 合格率统计
        pass_tests = result_query.filter_by(test_result='pass').count()
        pass_rate = (pass_tests / total_tests * 100) if total_tests > 0 else 0
        current_app.logger.info(f"合格率统计: 合格={pass_tests}, 合格率={pass_rate}%")

        # 平均阻抗值
        avg_rs = result_query.with_entities(func.avg(TestResult.rs_value)).scalar() or 0
        avg_rct = result_query.with_entities(func.avg(TestResult.rct_value)).scalar() or 0
        current_app.logger.info(f"平均阻抗值: Rs={avg_rs}, Rct={avg_rct}")

        result = {
            'total_devices': total_devices,
            'online_devices': online_devices,
            'total_tests': total_tests,
            'today_tests': today_tests,
            'pass_rate': round(pass_rate, 1),
            'avg_rs': round(float(avg_rs), 2),
            'avg_rct': round(float(avg_rct), 2),
            'status_code': 200
        }

        current_app.logger.info(f"仪表板数据返回成功: {result}")
        return jsonify(result), 200

    except Exception as e:
        import traceback
        error_msg = f"获取仪表板统计数据失败: {str(e)}"
        current_app.logger.error(error_msg)
        current_app.logger.error(f"错误堆栈: {traceback.format_exc()}")
        return jsonify({'error': 'Internal server error', 'details': str(e)}), 500

@web_api_bp.route('/devices/management', methods=['GET'])
@jwt_required()
def get_devices_for_management():
    """获取设备管理页面数据"""
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # 获取查询参数
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 20, type=int), 100)
        status_filter = request.args.get('status')
        search = request.args.get('search')
        
        # 构建查询
        query = Device.query
        
        # 权限控制
        if user.role != 'admin':
            query = query.filter_by(user_id=current_user_id)
        
        # 状态过滤
        if status_filter and status_filter != 'all':
            query = query.filter_by(status=status_filter)
        
        # 搜索过滤
        if search:
            search_pattern = f'%{search}%'
            query = query.filter(or_(
                Device.name.ilike(search_pattern),
                Device.device_id.ilike(search_pattern),
                Device.location.ilike(search_pattern)
            ))
        
        # 分页查询
        pagination = query.paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        devices = []
        for device in pagination.items:
            # 获取设备的测试统计
            test_count = TestResult.query.join(TestBatch).filter(
                TestBatch.device_id == device.id
            ).count()
            
            pass_count = TestResult.query.join(TestBatch).filter(
                and_(
                    TestBatch.device_id == device.id,
                    TestResult.test_result == 'pass'
                )
            ).count()
            
            pass_rate = (pass_count / test_count * 100) if test_count > 0 else 0
            
            device_data = device.to_dict()
            device_data.update({
                'total_tests': test_count,
                'pass_rate': round(pass_rate, 1),
                'last_seen': device.updated_at.isoformat() if device.updated_at else None
            })
            devices.append(device_data)
        
        return jsonify({
            'devices': devices,
            'pagination': {
                'page': pagination.page,
                'pages': pagination.pages,
                'per_page': pagination.per_page,
                'total': pagination.total,
                'has_next': pagination.has_next,
                'has_prev': pagination.has_prev
            },
            'status_code': 200
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"获取设备管理数据失败: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@web_api_bp.route('/analysis/test-results', methods=['GET'])
@jwt_required()
def get_test_results_for_analysis():
    """获取数据分析页面的测试结果"""
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # 获取查询参数
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 50, type=int), 200)
        device_id = request.args.get('device_id')
        batch_id = request.args.get('batch_id')
        cell_type = request.args.get('cell_type')
        channels = request.args.getlist('channels')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        result_filter = request.args.get('result')
        search = request.args.get('search')
        
        # 构建查询
        query = TestResult.query.join(TestBatch).join(Device)
        
        # 权限控制
        if user.role != 'admin':
            query = query.filter(Device.user_id == current_user_id)
        
        # 设备过滤
        if device_id:
            query = query.filter(Device.device_id == device_id)
        
        # 批次过滤
        if batch_id:
            query = query.filter(TestBatch.batch_id == batch_id)
        
        # 电芯类型过滤
        if cell_type:
            query = query.join(Battery).filter(Battery.cell_type == cell_type)
        
        # 通道过滤
        if channels:
            channel_nums = [int(ch) for ch in channels if ch.isdigit()]
            if channel_nums:
                query = query.filter(TestResult.channel_number.in_(channel_nums))
        
        # 日期范围过滤
        if start_date:
            start_dt = datetime.fromisoformat(start_date)
            query = query.filter(TestResult.test_start_time >= start_dt)
        
        if end_date:
            end_dt = datetime.fromisoformat(end_date)
            query = query.filter(TestResult.test_start_time <= end_dt)
        
        # 结果过滤
        if result_filter and result_filter != 'all':
            query = query.filter(TestResult.test_result == result_filter)
        
        # 搜索过滤
        if search:
            search_pattern = f'%{search}%'
            query = query.filter(or_(
                Device.device_id.ilike(search_pattern),
                TestBatch.batch_id.ilike(search_pattern),
                TestResult.test_id.ilike(search_pattern)
            ))
        
        # 排序
        query = query.order_by(desc(TestResult.test_start_time))
        
        # 分页查询
        pagination = query.paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        results = []
        for result in pagination.items:
            result_data = {
                'id': result.id,
                'test_id': result.test_id,
                'device_id': result.test_batch.device.device_id,
                'batch_id': result.test_batch.batch_id,
                'channel': result.channel_number,
                'test_time': result.test_start_time.isoformat(),
                'voltage': float(result.voltage) if result.voltage else None,
                'rs': float(result.rs_value) if result.rs_value else None,
                'rct': float(result.rct_value) if result.rct_value else None,
                'rsei': float(result.rsei_value) if hasattr(result, 'rsei_value') and result.rsei_value else None,
                'temperature': float(result.temperature) if result.temperature else None,
                'result': result.test_result,
                'grade': getattr(result, 'grade', None),
                'cell_type': result.battery.cell_type if result.battery else None
            }
            results.append(result_data)
        
        return jsonify({
            'results': results,
            'pagination': {
                'page': pagination.page,
                'pages': pagination.pages,
                'per_page': pagination.per_page,
                'total': pagination.total,
                'has_next': pagination.has_next,
                'has_prev': pagination.has_prev
            },
            'status_code': 200
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"获取测试结果数据失败: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@web_api_bp.route('/analysis/filter-options', methods=['GET'])
@jwt_required()
def get_filter_options():
    """获取数据分析页面的筛选选项"""
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # 构建基础查询
        if user.role == 'admin':
            device_query = Device.query
            batch_query = TestBatch.query.join(Device)
            battery_query = Battery.query.join(TestResult).join(TestBatch).join(Device)
        else:
            device_query = Device.query.filter_by(user_id=current_user_id)
            batch_query = TestBatch.query.join(Device).filter(Device.user_id == current_user_id)
            battery_query = Battery.query.join(TestResult).join(TestBatch).join(Device).filter(Device.user_id == current_user_id)
        
        # 获取设备列表
        devices = device_query.with_entities(Device.device_id).distinct().all()
        device_list = [d.device_id for d in devices]
        
        # 获取批次列表
        batches = batch_query.with_entities(TestBatch.batch_id).distinct().all()
        batch_list = [b.batch_id for b in batches]
        
        # 获取电芯类型列表
        cell_types = battery_query.with_entities(Battery.cell_type).distinct().all()
        cell_type_list = [ct.cell_type for ct in cell_types if ct.cell_type]
        
        return jsonify({
            'devices': device_list,
            'batches': batch_list,
            'cell_types': cell_type_list,
            'channels': list(range(1, 9)),  # 1-8通道
            'status_code': 200
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"获取筛选选项失败: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@web_api_bp.route('/analysis/impedance-details/<int:test_result_id>', methods=['GET'])
@jwt_required()
def get_impedance_details(test_result_id):
    """获取测试结果的阻抗明细数据"""
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # 获取测试结果
        test_result = TestResult.query.get(test_result_id)
        if not test_result:
            return jsonify({'error': 'Test result not found'}), 404
        
        # 权限检查
        if user.role != 'admin' and test_result.test_batch.device.user_id != current_user_id:
            return jsonify({'error': 'Access denied'}), 403
        
        # 获取阻抗明细数据
        impedance_details = ImpedanceDetail.query.filter_by(
            test_id=test_result_id
        ).order_by(ImpedanceDetail.frequency).all()

        details = []
        for detail in impedance_details:
            detail_data = {
                'id': detail.id,
                'frequency': float(detail.frequency) if detail.frequency else None,
                'z_real': float(detail.z_real) if detail.z_real else None,
                'z_imag': float(detail.z_imag) if detail.z_imag else None,
                'z_magnitude': float(detail.z_magnitude) if detail.z_magnitude else None,
                'phase_angle': float(detail.phase_angle) if detail.phase_angle else None,
                'measurement_time': detail.measurement_time.isoformat() if detail.measurement_time else None
            }
            details.append(detail_data)

        return jsonify({
            'test_result_id': test_result_id,
            'details': details,
            'status_code': 200
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"获取阻抗明细数据失败: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@web_api_bp.route('/recent-tests', methods=['GET'])
@jwt_required()
def get_recent_tests():
    """获取最近的测试数据"""
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        limit = min(request.args.get('limit', 10, type=int), 50)
        
        # 构建查询
        query = TestResult.query.join(TestBatch).join(Device)
        
        # 权限控制
        if user.role != 'admin':
            query = query.filter(Device.user_id == current_user_id)
        
        # 获取最近的测试
        recent_tests = query.order_by(desc(TestResult.test_start_time)).limit(limit).all()
        
        results = []
        for test in recent_tests:
            result_data = {
                'id': test.id,
                'device_id': test.test_batch.device.device_id,
                'channel': test.channel_number,
                'test_time': test.test_start_time.isoformat(),
                'result': test.test_result,
                'rs': float(test.rs_value) if test.rs_value else None,
                'rct': float(test.rct_value) if test.rct_value else None
            }
            results.append(result_data)
        
        return jsonify({
            'recent_tests': results,
            'status_code': 200
        }), 200

    except Exception as e:
        current_app.logger.error(f"获取最近测试数据失败: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@web_api_bp.route('/export/test-results', methods=['POST'])
@jwt_required()
def export_test_results():
    """导出测试结果数据"""
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)

        if not user:
            return jsonify({'error': 'User not found'}), 404

        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        test_result_ids = data.get('test_result_ids', [])
        export_format = data.get('format', 'csv').lower()
        selected_fields = data.get('fields', [])

        if not test_result_ids:
            return jsonify({'error': 'No test results selected'}), 400

        # 获取测试结果数据
        query = TestResult.query.join(TestBatch).join(Device).filter(
            TestResult.id.in_(test_result_ids)
        )

        # 权限控制
        if user.role != 'admin':
            query = query.filter(Device.user_id == current_user_id)

        test_results = query.all()

        if not test_results:
            return jsonify({'error': 'No accessible test results found'}), 404

        # 准备导出数据
        export_data = []
        for result in test_results:
            row_data = {
                'ID': result.id,
                '测试ID': result.test_id,
                '设备ID': result.test_batch.device.device_id,
                '批次ID': result.test_batch.batch_id,
                '通道': result.channel_number,
                '测试时间': result.test_start_time.strftime('%Y-%m-%d %H:%M:%S'),
                '电压(V)': float(result.voltage) if result.voltage else None,
                'Rs(mΩ)': float(result.rs_value) if result.rs_value else None,
                'Rct(mΩ)': float(result.rct_value) if result.rct_value else None,
                '温度(°C)': float(result.temperature) if result.temperature else None,
                '测试结果': result.test_result,
                '错误代码': result.error_code,
                '电芯类型': result.battery.cell_type if result.battery else None,
                '创建时间': result.created_at.strftime('%Y-%m-%d %H:%M:%S') if result.created_at else None
            }

            # 如果指定了字段，只导出选中的字段
            if selected_fields:
                filtered_data = {k: v for k, v in row_data.items() if k in selected_fields}
                export_data.append(filtered_data)
            else:
                export_data.append(row_data)

        # 根据格式导出
        if export_format == 'excel' and EXCEL_SUPPORT:
            return _export_to_excel(export_data, f'test_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        else:
            return _export_to_csv(export_data, f'test_results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv')

    except Exception as e:
        current_app.logger.error(f"导出测试结果失败: {str(e)}")
        return jsonify({'error': 'Export failed'}), 500

def _export_to_csv(data, filename):
    """导出为CSV格式"""
    if not data:
        return jsonify({'error': 'No data to export'}), 400

    # 创建CSV内容
    output = io.StringIO()

    # 获取字段名
    fieldnames = list(data[0].keys())
    writer = csv.DictWriter(output, fieldnames=fieldnames)

    # 写入表头
    writer.writeheader()

    # 写入数据
    for row in data:
        # 处理None值
        processed_row = {k: (v if v is not None else '') for k, v in row.items()}
        writer.writerow(processed_row)

    # 创建响应
    output.seek(0)
    csv_data = output.getvalue()
    output.close()

    # 创建临时文件
    temp_file = tempfile.NamedTemporaryFile(mode='w+', delete=False, suffix='.csv', encoding='utf-8-sig')
    temp_file.write(csv_data)
    temp_file.close()

    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name=filename,
        mimetype='text/csv'
    )

def _export_to_excel(data, filename):
    """导出为Excel格式"""
    if not EXCEL_SUPPORT:
        return jsonify({'error': 'Excel export not supported'}), 400

    if not data:
        return jsonify({'error': 'No data to export'}), 400

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "测试结果"

    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 获取字段名并写入表头
    fieldnames = list(data[0].keys())
    for col, fieldname in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col, value=fieldname)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, fieldname in enumerate(fieldnames, 1):
            value = row_data.get(fieldname)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 保存到临时文件
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()

    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@web_api_bp.route('/export/impedance-details', methods=['POST'])
@jwt_required()
def export_impedance_details():
    """导出阻抗明细数据"""
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)

        if not user:
            return jsonify({'error': 'User not found'}), 404

        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        test_result_ids = data.get('test_result_ids', [])
        export_format = data.get('format', 'csv').lower()

        if not test_result_ids:
            return jsonify({'error': 'No test results selected'}), 400

        # 获取阻抗明细数据
        query = ImpedanceDetail.query.join(TestResult).join(TestBatch).join(Device).filter(
            ImpedanceDetail.test_result_id.in_(test_result_ids)
        )

        # 权限控制
        if user.role != 'admin':
            query = query.filter(Device.user_id == current_user_id)

        impedance_details = query.order_by(
            ImpedanceDetail.test_result_id,
            ImpedanceDetail.frequency
        ).all()

        if not impedance_details:
            return jsonify({'error': 'No accessible impedance details found'}), 404

        # 准备导出数据
        export_data = []
        for detail in impedance_details:
            row_data = {
                '测试结果ID': detail.test_result_id,
                '测试ID': detail.test_result.test_id,
                '设备ID': detail.test_result.test_batch.device.device_id,
                '通道': detail.test_result.channel_number,
                '频率(Hz)': float(detail.frequency),
                '实部阻抗(mΩ)': float(detail.real_impedance),
                '虚部阻抗(mΩ)': float(detail.imaginary_impedance),
                '阻抗模值(mΩ)': float(detail.magnitude),
                '相位角(°)': float(detail.phase),
                '测试时间': detail.test_result.test_start_time.strftime('%Y-%m-%d %H:%M:%S')
            }
            export_data.append(row_data)

        # 根据格式导出
        if export_format == 'excel' and EXCEL_SUPPORT:
            return _export_to_excel(export_data, f'impedance_details_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        else:
            return _export_to_csv(export_data, f'impedance_details_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv')

    except Exception as e:
        current_app.logger.error(f"导出阻抗明细数据失败: {str(e)}")
        return jsonify({'error': 'Export failed'}), 500
